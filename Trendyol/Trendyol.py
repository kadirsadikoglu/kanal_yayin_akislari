import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from time import sleep
from bs4 import BeautifulSoup
import openpyxl
import xlsxwriter
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options

workbook = xlsxwriter.Workbook('kupon_kontrol.xlsx')
worksheet = workbook.add_worksheet("Kupon Kontrol")
workbook.formats[0].set_font_size(8)

data = openpyxl.load_workbook("D:\pythonProject\Trendyol\Trendyol Hesap Listesi.xlsx")
s = data.active

def _options():
    options = Options()
    options.add_argument('--ignore-certificate-errors')
    #options.add_argument("--test-type")
    options.add_argument("--headless")
    options.add_argument("--incognito")
    options.add_argument('--disable-gpu') if os.name == 'nt' else None # Windows workaround
    options.add_argument("--verbose")
    return options

def kupon_arama(x,y,z):
    chromedriver = 'C:/Users/kadir/chromedriver.exe'
    service1 = Service(chromedriver)
    driver = webdriver.Chrome(service=service1)
    URL = 'https://www.trendyol.com/giris'
    driver.get(URL)
    sleep(3)
    driver.find_element(By.XPATH, "//*[@id='login-email']").send_keys(x)
    sleep(1)
    driver.find_element(By.XPATH, "//*[@id='login-password-input']").send_keys(y)
    sleep(1)
    driver.find_element(By.XPATH, "//*[@class='q-primary q-fluid q-button-medium q-button submit']").click()
    kupon_sayfasi = "https://www.trendyol.com/Hesabim/IndirimKuponlari"
    sleep(3)
    driver.get(kupon_sayfasi)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    sleep(3)
    kupon_ara = soup.find_all('span', attrs={'class': 'coupon-name'})
    for i in kupon_ara:
        print(i.get_text())
        worksheet.write(abc, 2, i.get_text())
    driver.quit()


for i in range(2,200):
    abc = i
    mail1 = s.cell(row=i, column=1)
    password1 = s.cell(row=i, column=2)
    kupon_arama(mail1.value, password1.value,abc)
    worksheet.write(i, 0, mail1.value)
    worksheet.write(i, 1, password1.value)

workbook.close()
